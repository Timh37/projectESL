#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Nov 30 09:14:59 2023

@author: timhermans
"""

import geopandas as geopd
from shapely.geometry import Point
from utils import angdist
import numpy as np
from tqdm import tqdm
import os
from openpyxl import load_workbook
import pandas as pd

def find_flopros_protection_levels(qlons,qlats,flopros_dir,maxdist):

    polygons = geopd.read_file(os.path.join(flopros_dir,'Results_adaptation_objectives/Countries_States_simplified.shp')) #shape file with region polygons
    wb = load_workbook(filename = os.path.join(flopros_dir,'FLOPROS_geogunit_107.xlsx')) #protection standards for each region
    ws = wb.active
    flopros = np.array([cell.value for cell in ws['D'][1::]],dtype=float)
    
    centroids = polygons.centroid #determine polygon centroids for quick filtering
    latc = [k.coords.xy[-1][0] for k in centroids]
    lonc = [k.coords.xy[0][0] for k in centroids]
    
    nearest_segments = []
    
    polygons = polygons.set_crs('epsg:4326')
    
    for qlat,qlon in tqdm(zip(qlats,qlons)): #loop over query coordinates
        p = Point(qlon,qlat) #put coords into point geometry
        
        #do a first filtering based on angular distance to segment/polygon center points (more efficient than shapely distance)
        angdists = angdist(qlat,qlon,latc,lonc)
        nearby = np.where(angdists<5)[0]
        if len(nearby)==0:
            nearest_segments.append(np.nan)
            #add a warning?
            continue
        
        i_near = polygons.iloc[nearby].index[np.argsort( [p.distance(k) for k in polygons.iloc[nearby].geometry])[0:3]] #find the nearest 3 segments based on euclidean distances (default for shapely)
        
        #for these segments, refine the distance computation using a coordinate-specific appropriate reference system
        p_gdf = geopd.GeoDataFrame(geometry=[p], crs='epsg:4326') #put query point into cartesian geodataframe
        if qlat<0: #degermine appropriate epsg system
            e = 32701 + np.floor_divide((qlon+180),6)
            if qlat<-80:
                e = 32761
        else:
            e = 32601 + np.floor_divide((qlon+180),6)
            if qlat>80:
                e = 32661
            
        p_gdf.to_crs(epsg=e, inplace=True) #convert point to epsg system
        polygons_ = polygons.iloc[i_near].to_crs(epsg=e) #convert coordinates of 3 nearest diva segments to epsg system
        
        km_dists = np.array([p_gdf.distance(k).values[0] for k in polygons_.geometry])/1000 #compute kilometric distances
        
        if np.min(km_dists>maxdist): #if nearest point on nearest segment is more than 15km away
            nearest_segments.append(np.nan)
            #add a warning?
            continue
    
        nearest_segments.append(polygons_.index[np.argmin( km_dists )]) #append index of nearest diva segment
        
    protection_levels = []
    
    for k in nearest_segments:
        if np.isfinite(k):
            plevel = flopros[polygons.FID_Aque.iloc[k]] #in units of return period
            protection_levels.append(1/plevel) #add protection level to list in units of return frequency
        else:
            protection_levels.append(np.nan)
            
    return np.array(protection_levels)



if __name__ == "__main__":
    maxdist=15
    g3_meta = pd.read_csv('/Volumes/Naamloos/PhD_Data/GESLA3/GESLA3_ALL.csv')
    qlats = g3_meta['LATITUDE'].values
    qlons = g3_meta['LONGITUDE'].values
    
    flopros_dir = '/Users/timhermans/Documents/Data/FLOPROS/Tiggeloven/'
    protection_levels = find_flopros_protection_levels(qlons,qlats,flopros_dir,maxdist)
    
    polygons = geopd.read_file(os.path.join(flopros_dir,'Results_adaptation_objectives/Countries_States_simplified.shp')) #shape file with region polygons
    wb = load_workbook(filename = os.path.join(flopros_dir,'FLOPROS_geogunit_107.xlsx')) #protection standards for each region
    ws = wb.active
    flopros = np.array([cell.value for cell in ws['D'][1::]],dtype=float)

    polygons['FLOPROS'] = [flopros[k] for k in polygons['FID_Aque'].values]
    
    import matplotlib.pyplot as plt
    from cartopy import crs as ccrs
    
    fig=plt.figure(figsize=(8,8)) #generate figure  
    gs = fig.add_gridspec(1,1)
    crs = ccrs.Robinson(central_longitude=0)
    
    ax = plt.subplot(gs[0,0],projection=ccrs.Robinson(central_longitude=0))
    
    # This can be converted into a `proj4` string/dict compatible with GeoPandas
    crs_proj4 = crs.proj4_init
    df_ae = polygons.to_crs(crs_proj4)
    df_ae.plot(column='FLOPROS',vmin=0,vmax=100,cmap='Reds',ax=ax)
    ax.scatter(qlons[np.isfinite(protection_levels)],qlats[np.isfinite(protection_levels)],c=1/np.array(protection_levels)[np.isfinite(protection_levels)],vmin=0,vmax=100,s=10,cmap='Reds',edgecolors='black',transform=ccrs.PlateCarree(),zorder=5)