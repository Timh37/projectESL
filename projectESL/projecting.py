'''
@author: Tim Hermans
t(dot)h(dot)j(dot)hermans@uu(dot)nl
'''
import numpy as np
import geopandas as geopd
from shapely.geometry import Point
from utils import angdist
from tqdm import tqdm
import os
from openpyxl import load_workbook

def compute_AF(f,z_hist,slr,refFreq):
    '''computes amplification factor based on historical return curve(s) "z_hist" = F("f"), sea-level projections "slr" in a given year and reference frequency "refFreq"'''
    i_ref = np.argmin(np.abs(f-refFreq)) #find index of f closest to refFreq
    i_z_min = np.nanargmin(z_hist,axis=0)[0]
    
    if (z_hist.ndim == 1): #if no z_hist samples
        if np.isscalar(slr)==False:
            z_hist = np.repeat(z_hist[:,None],len(slr),axis=1)
    
    z_fut = z_hist+slr
        
    refZ_hist = z_hist[i_ref] #historical reference height samples corresponding to refFreq
    
    iRefZ_hist_in_z_fut = np.nanargmin(np.abs(z_fut - refZ_hist),axis=0) #find future frequency corresponding to those heights
    
    AF = f[iRefZ_hist_in_z_fut]/refFreq #divide future frequency correspondong to refZ_hist by historical reference frequency

    max_AF = f[i_z_min]/f[i_ref] #keep track of what AF could maximally be given defined z
    
    return z_fut,AF,max_AF

def compute_AF_timing(f,z_hist,slr,refFreq,AF):
    '''computes amplification factor timing based on historical return curve(s) "z_hist" = F("f"), sea-level projections "slr" as function of years and reference frequency "refFreq"'''
    if AF<=1:
        timing = np.zeros(len(slr.samples)) -1
        print("Warning: AF>=1; returning projected timing = -1.")
        return timing
    
    i_ref = np.argmin(np.abs(f-refFreq)) #find index of f closest to refFreq
    i_ref_AF = np.argmin(np.abs(f-refFreq*AF)) #find index of f closest to AF * refFreq

    if (z_hist.ndim == 1):
        z_hist = np.repeat(z_hist[:,None],len(slr.samples),axis=1)
 
    req_slr = z_hist[i_ref] - z_hist[i_ref_AF] #sea-level rise required to go from refFreq to AF*refFreq
    
    #find first year in which SLR > required SLR
    slr_minus_required = slr.values - np.repeat(req_slr[:,np.newaxis],len(slr.years),axis=1)
    slr_minus_required[slr_minus_required<0] = 999 #set AFs for which sea-level fall is required to 999
    
    try:
        imin = np.nanargmin(slr_minus_required,axis=-1)
        timing = slr.years.values[imin]
    except: #if encountering all nan slice, req_slr is nan because i_ref/i_ref_AF is not supported in z_hist
        timing = np.zeros(len(slr.samples)) - 1
        print("Warning: Could not compute required SLR from historical return curve for refFreq={0} and AF={1}; returning projected timing = -1.".format(refFreq,AF))
        return timing
    
    try:
        timing[timing==slr.years.values[-1]] = np.nan #end of timeseries or later -> cannot evaluate timing, so set to nan
    except:
        pass
    
    return timing


def find_flopros_protection_levels(qlons,qlats,flopros_dir,maxdist):
    '''find flood protection level of polygons nearest (within "maxdist") to queried sites (qlons,qlats)'''
    polygons = geopd.read_file(os.path.join(flopros_dir,'Results_adaptation_objectives/Countries_States_simplified.shp')) #shape file with region polygons
    wb = load_workbook(filename = os.path.join(flopros_dir,'FLOPROS_geogunit_107.xlsx')) #protection standards for each region
    ws = wb.active
    flopros = np.array([cell.value for cell in ws['D'][1::]],dtype=float)
    
    centroids = polygons.centroid #determine polygon centroids for quick filtering
    latc = [k.coords.xy[-1][0] for k in centroids]
    lonc = [k.coords.xy[0][0] for k in centroids]
    
    nearest_segments = []
    
    polygons = polygons.set_crs('epsg:4326')
    
    for qlat,qlon in tqdm(zip(qlats,qlons)): #loop over query coordinates
        p = Point(qlon,qlat) #put coords into point geometry
        
        #do a first filtering based on angular distance to segment/polygon center points (more efficient than shapely distance)
        angdists = angdist(qlat,qlon,latc,lonc)
        nearby = np.where(angdists<5)[0]
        if len(nearby)==0:
            nearest_segments.append(np.nan)
            #add a warning?
            continue
        
        i_near = polygons.iloc[nearby].index[np.argsort( [p.distance(k) for k in polygons.iloc[nearby].geometry])[0:3]] #find the nearest 3 segments based on euclidean distances (default for shapely)
        
        #for these segments, refine the distance computation using a coordinate-specific appropriate reference system
        p_gdf = geopd.GeoDataFrame(geometry=[p], crs='epsg:4326') #put query point into cartesian geodataframe
        if qlat<0: #degermine appropriate epsg system
            e = 32701 + np.floor_divide((qlon+180),6)
            if qlat<-80:
                e = 32761
        else:
            e = 32601 + np.floor_divide((qlon+180),6)
            if qlat>80:
                e = 32661
            
        p_gdf.to_crs(epsg=e, inplace=True) #convert point to epsg system
        polygons_ = polygons.iloc[i_near].to_crs(epsg=e) #convert coordinates of 3 nearest diva segments to epsg system
        
        km_dists = np.array([p_gdf.distance(k).values[0] for k in polygons_.geometry])/1000 #compute kilometric distances
        
        if np.min(km_dists>maxdist): #if nearest point on nearest segment is more than 15km away
            nearest_segments.append(np.nan)
            #add a warning?
            continue
    
        nearest_segments.append(polygons_.index[np.argmin( km_dists )]) #append index of nearest diva segment
        
    protection_levels = []
    
    for k in nearest_segments:
        if np.isfinite(k):
            plevel = flopros[polygons.FID_Aque.iloc[k]] #in units of return period
            protection_levels.append(1/plevel) #add protection level to list in units of return frequency
        else:
            protection_levels.append(np.nan)
            
    return np.array(protection_levels)


def find_diva_protection_levels(qlons,qlats,diva_fn,maxdist):
    '''find flood protection level of polygons nearest (within "maxdist") to queried sites (qlons,qlats)'''
    diva = geopd.read_file(diva_fn) #open diva geo file
    
    nearest_segments = []
    
    for qlat,qlon in tqdm(zip(qlats,qlons)): #loop over query coordinates
        p = Point(qlon,qlat) #put coords into point geometry
        
        #do a first filtering based on angular distance to segment/polygon center points (more efficient than shapely distance)
        angdists = angdist(qlat,qlon,diva.lati.values,diva.longi.values)
        nearby = np.where(angdists<5)[0]
        if len(nearby)==0:
            nearest_segments.append(np.nan)
            #add a warning?
            continue
        
        i_near = diva.iloc[nearby].index[np.argsort( [p.distance(k) for k in diva.iloc[nearby].geometry])[0:5]] #find the nearest 5 segments based on euclidean distances (default for shapely)
        
        #for these segments, refine the distance computation using a coordinate-specific appropriate reference system
        p_gdf = geopd.GeoDataFrame(geometry=[p], crs='epsg:4326') #put query point into cartesian geodataframe
        if qlat<0: #degermine appropriate epsg system
            e = 32701 + np.floor_divide((qlon+180),6)
            if qlat<-80:
                e = 32761
        else:
            e = 32601 + np.floor_divide((qlon+180),6)
            if qlat>80:
                e = 32661
            
        p_gdf.to_crs(epsg=e, inplace=True) #convert point to epsg system
        diva_ = diva.iloc[i_near].to_crs(epsg=e) #convert coordinates of 5 nearest diva segments to epsg system
        
        km_dists = np.array([p_gdf.distance(k).values[0] for k in diva_.geometry])/1000 #compute kilometric distances
        
        if np.min(km_dists>maxdist): #if nearest point on nearest segment is more than 15km away
            nearest_segments.append(np.nan)
            #add a warning?
            continue
    
        nearest_segments.append(diva_.index[np.argmin( km_dists )]) #append index of nearest diva segment
    
    protection_levels = []

    for k in nearest_segments:
        if np.isfinite(k):
            plevel = diva.protection_level_modelled.values[k] #in units of return period
            
            if plevel==0: #if no protection, assign protection of 1/2y (see Scussolini et al., 2016)
                plevel = 2
            protection_levels.append(1/plevel) #add protection level to list in units of return frequency
        else:
            protection_levels.append(np.nan)
    return np.array(protection_levels)